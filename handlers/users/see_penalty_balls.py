import os
import tempfile

from aiogram import types
from aiogram.dispatcher import FSMContext

from keyboards.default.go_to_registration import go_registration_default_keyboard
from keyboards.default.menu_keyboards import back_to_menu
from keyboards.default.period_keyboards import period_default_keyboard
from loader import dp, db, bot
from states.penaltyball import GetPenaltyBallState
from openpyxl.styles import Alignment
import openpyxl


async def download_penalty_balls(days):
    students = await db.select_all_students()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = "F.I.Sh"
    worksheet['C1'] = 'Jazo bali'

    for cell in ['A1', 'B1', 'C1']:
        worksheet[cell].alignment = Alignment(horizontal='center')

    worksheet.cell(row=1, column=1, value='T/r')
    worksheet.cell(row=1, column=2, value='F.I.Sh')
    worksheet.cell(row=1, column=3, value="Jazo bali")

    penalty_balls_dict = dict()
    for student in students:
        full_name = f"{student['first_name']} {student['last_name']}"
        total_penalty_ball = 0
        if days == "all":
            penalty_balls = await db.select_penalty_balls(student_id=student['id'])
        else:
            penalty_balls = await db.select_penalty_ball_by_period(student_id=student['id'], days=days)
        for penalty_ball in penalty_balls:
            total_penalty_ball += penalty_ball['ball']
        if total_penalty_ball > 0:
            penalty_balls_dict[full_name] = total_penalty_ball

    sorted_penalty_balls = sorted(penalty_balls_dict.items(), key=lambda item: item[1], reverse=True)
    tr = 0
    for student_ball in sorted_penalty_balls:
        student = student_ball[0]
        ball = student_ball[1]
        tr += 1
        worksheet.cell(row=tr + 1, column=1, value=tr).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=2, value=student).alignment = Alignment(horizontal='center')
        worksheet.cell(row=tr + 1, column=3, value=ball).alignment = Alignment(horizontal='center')

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'PenaltyBalls.xlsx')
    workbook.save(file_path)

    return temp_dir


@dp.message_handler(text="📊 Jarima ballarini ko'rish", state="*")
async def delete_teacher(message: types.Message, state: FSMContext):
    try:
        await state.finish()
    except:
        pass
    user_telegram_id = message.from_user.id
    users = await db.select_users(telegram_id=user_telegram_id)
    if not users:
        await message.answer(text="🚫 Sizda botdan foydalanish uchun ruxsat mavjud emas,\n"
                                  "📝 Botdan foydalanish uchun ro'yxatdan o'tishingiz kerak 👇",
                             reply_markup=go_registration_default_keyboard)
    else:
        user = users[0]
        user_role = user['role']
        if user_role != 'admin':
            await message.reply(text="⚠️ Bu buyruqdan foydalanish uchun sizda ruxsat mavjud emas!",
                                reply_markup=back_to_menu)
            return
        await message.answer(text="Baholarni ko'rish uchun muddatini tanlang 👇", reply_markup=period_default_keyboard)
        await GetPenaltyBallState.period.set()


@dp.message_handler(state=GetPenaltyBallState.period)
async def get_period_function(message: types.Message, state: FSMContext):
    period = message.text
    if period == "1 kun":
        days = 1
    elif period == "5 kun":
        days = 5
    elif period == "1 hafta":
        days = 7
    elif period == "10 kun":
        days = 10
    elif period == "15 kun":
        days = 15
    elif period == "20 kun":
        days = 20
    elif period == "1 oy":
        days = 30
    elif period == "2 oy":
        days = 60
    elif period == "3 oy":
        days = 90
    elif period == "6 oy":
        days = 180
    elif period == "1 yil":
        days = 365
    elif period == "Hammasi":
        days = "all"
    else:
        await message.answer(text="Siz quyidagi muddatlardan birini tanlashingiz kerak 👇",
                             reply_markup=period_default_keyboard)
        return
    temp_dir = await download_penalty_balls(days=days)

    with open(os.path.join(temp_dir, 'PenaltyBalls.xlsx'), 'rb') as file:
        await message.answer_document(document=file)

    os.remove(os.path.join(temp_dir, 'PenaltyBalls.xlsx'))
